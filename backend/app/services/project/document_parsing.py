"""Document parsing helpers for Project Management.

Single responsibility: convert uploaded file bytes into normalized, human-readable
text suitable for downstream requirement extraction.

Parsing failures return (None, reason) rather than raising, so callers can record
partial success (FR-001, SC-004).
"""

from __future__ import annotations

import logging
from io import BytesIO

import openpyxl
import xlrd
from pypdf import PdfReader

logger = logging.getLogger(__name__)


def extract_text_from_upload(
    *,
    file_name: str,
    mime_type: str | None,
    content: bytes,
) -> tuple[str | None, str | None]:
    """Extract human-readable text from an uploaded file.

    Returns:
        (text, None) on success
        (None, reason) on failure
    """
    safe_name = (file_name or "").strip()
    lower_name = safe_name.lower()

    if lower_name.endswith(".pdf") or (mime_type or "") == "application/pdf":
        return _extract_pdf_text(content)

    if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        return _extract_excel_text(lower_name, content)

    # Default: treat as text
    try:
        return content.decode("utf-8"), None
    except UnicodeDecodeError:
        # Keep going with a lossy decode rather than failing ingestion completely.
        return content.decode("utf-8", errors="replace"), None


def _extract_pdf_text(content: bytes) -> tuple[str | None, str | None]:
    try:
        reader = PdfReader(BytesIO(content))
        page_texts = []
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                page_texts.append(text)

        combined = "\n\n".join(page_texts).strip()
        if not combined:
            return None, "PDF contains no extractable text"
        return combined, None
    except Exception as exc:
        logger.exception("PDF extraction failed")
        return None, f"PDF extraction failed: {exc}"


def _extract_excel_text(
    lower_name: str, content: bytes
) -> tuple[str | None, str | None]:
    if lower_name.endswith(".xlsx"):
        return _extract_xlsx_text(content)
    return _extract_xls_text(content)


def _extract_xlsx_text(content: bytes) -> tuple[str | None, str | None]:
    try:
        workbook = openpyxl.load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )

        lines = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"# Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                if any(v.strip() for v in values):
                    lines.append("\t".join(values))

        text = "\n".join(lines).strip()
        if not text:
            return None, "XLSX contains no extractable text"
        return text, None
    except Exception as exc:
        logger.exception("XLSX extraction failed")
        return None, f"XLSX extraction failed: {exc}"


def _extract_xls_text(content: bytes) -> tuple[str | None, str | None]:
    try:
        book = xlrd.open_workbook(file_contents=content)
        lines = []
        for sheet in book.sheets():
            lines.append(f"# Sheet: {sheet.name}")
            for r in range(sheet.nrows):
                values = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                if any(v.strip() for v in values):
                    lines.append("\t".join(values))

        text = "\n".join(lines).strip()
        if not text:
            return None, "XLS contains no extractable text"
        return text, None
    except Exception as exc:
        logger.exception("XLS extraction failed")
        return None, f"XLS extraction failed: {exc}"

